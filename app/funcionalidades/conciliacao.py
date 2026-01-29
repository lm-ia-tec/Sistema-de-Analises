import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from io import BytesIO
import utils  # Importando nosso arquivo de utilitários
import pickle
import os
import io
import hashlib

# =========================================================
# PÁGINA — CONCILIAÇÃO ISS (INTERFACE ORIGINAL)
# =========================================================
def carregar_arquivo_csv(arquivo, sep=None, decimal=None, **kwargs):
    # Prioridade para ponto e vírgula
    candidatos_sep = [sep, ';', ',', '\t', None]
    candidatos_encoding = [kwargs.pop('encoding', None), 'utf-8', 'latin-1', 'utf-16']
    candidatos_decimal = [decimal, ',', '.']
    for s in candidatos_sep:
        for enc in candidatos_encoding:
            for dec in candidatos_decimal:
                try:
                    params = dict(sep=s, encoding=enc, decimal=dec, engine='python', **kwargs)
                    if hasattr(arquivo, "seek"):
                        arquivo.seek(0)
                    df = pd.read_csv(arquivo, **{k: v for k, v in params.items() if v is not None})
                    if isinstance(df, pd.DataFrame) and df.shape[1] >= 1:
                        return df
                except Exception:
                    continue
    return pd.DataFrame()

def format_cnpj(cnpj):
    cnpj = str(cnpj).replace('.', '').replace('/', '').replace('-', '').replace(' ', '')
    if len(cnpj) < 14 and cnpj.isdigit():
        cnpj = cnpj.zfill(14)
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj

def parse_moeda_brasil_robusto(serie):
    s = (serie.astype(str)
                 .str.replace(r'[^0-9,.\-]', '', regex=True)
                 .str.replace('.', '', regex=False)
                 .str.replace(',', '.', regex=False))
    return pd.to_numeric(s, errors='coerce')

def preparar_dataframe_fortaleza(file_like):
    try:
        xls = pd.ExcelFile(file_like)
    except Exception:
        return pd.DataFrame()
    try:
        df_tomados = xls.parse('Serviços Tomados', header=0)
    except Exception:
        try:
            df_tomados = xls.parse(xls.sheet_names[0], header=0)
        except Exception:
            return pd.DataFrame()
    if 'Status Doc.' in df_tomados.columns:
        df_tomados = df_tomados[df_tomados['Status Doc.'] != 'CANCELADA']
    if 'ISS Retido' in df_tomados.columns:
        df_tomados = df_tomados[~df_tomados['ISS Retido'].isin(['Não', 'NÃO'])]
    try:
        df_pendentes = xls.parse('Serviços Pendentes', header=8)
        df_pendentes['Status Aceite'] = 'Pendente'
    except Exception:
        df_pendentes = pd.DataFrame()
    columns_tomados = ['Data', 'CPF/CNPJ Prestador', 'Razão Social/Nome do Prestador',
                       'Número', 'Valor do ISS', 'Valor dos Serviços', 'ISS Retido', 'Status Aceite']
    columns_pendentes = ['Data', 'CNPJ/CPF Prestador', 'Razão Social/Nome do Prestador',
                          'Número', 'Valor do ISS', 'Valor do Serviço', 'ISS Retido', 'Status Aceite']
    df_tomados = df_tomados[[c for c in columns_tomados if c in df_tomados.columns]].copy()
    if not df_pendentes.empty:
        df_pendentes = df_pendentes[[c for c in columns_pendentes if c in df_pendentes.columns]].copy()
        df_pendentes = df_pendentes.rename(columns={
            'CNPJ/CPF Prestador': 'CPF/CNPJ Prestador',
            'Valor do Serviço': 'Valor dos Serviços'
        })
        merged_df = pd.concat([df_tomados, df_pendentes], ignore_index=True)
    else:
        merged_df = df_tomados.copy()
    merged_df['Origem'] = 'Fortaleza'
    if 'Status Doc.' not in merged_df.columns:
        merged_df['Status Doc.'] = None
    if 'CPF/CNPJ Prestador' in merged_df.columns:
        merged_df['CPF/CNPJ Prestador'] = merged_df['CPF/CNPJ Prestador'].astype(str)
    if 'Número' in merged_df.columns:
        merged_df['Número'] = merged_df['Número'].astype(str).str.replace(r'\.0$', '', regex=True)
    if 'Valor do ISS' in merged_df.columns:
        merged_df['Valor do ISS'] = pd.to_numeric(merged_df['Valor do ISS'], errors='coerce')
    if 'Valor dos Serviços' in merged_df.columns:
        merged_df['Valor dos Serviços'] = pd.to_numeric(merged_df['Valor dos Serviços'], errors='coerce')
    return merged_df

def preparar_dataframe_vr(file_like):
    try:
        df = pd.read_excel(file_like, skiprows=16)
    except Exception:
        try:
            file_like.seek(0)
            df = pd.read_excel(file_like)
        except Exception:
            return pd.DataFrame()
    rename_map = {
        'CNPJ Prestador': 'CPF/CNPJ Prestador',
        'Razão Social': 'Razão Social/Nome do Prestador',
        'Nº': 'Número',
        'Dt Emiss': 'Data',
        'Nota Fiscal': 'Valor dos Serviços',
        'Imposto': 'Valor do ISS',
        'Retido': 'ISS Retido',
        'Status': 'Status Doc.'
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns}).copy()
    if 'Razão Social/Nome do Prestador' in df.columns:
        df = df.dropna(subset=['Razão Social/Nome do Prestador'])
    if 'CPF/CNPJ Prestador' in df.columns:
        df['CPF/CNPJ Prestador'] = df['CPF/CNPJ Prestador'].astype(str)
    if 'Número' in df.columns:
        df['Número'] = df['Número'].astype(str).str.replace(r'\.0$', '', regex=True)
    df['Origem'] = 'Volta Redonda'
    if 'Valor do ISS' in df.columns:
        df['Valor do ISS'] = pd.to_numeric(df['Valor do ISS'], errors='coerce')
    if 'Valor dos Serviços' in df.columns:
        df['Valor dos Serviços'] = pd.to_numeric(df['Valor dos Serviços'], errors='coerce')
    return df

def unificar_dataframes(df1, df2):
    if (df1 is None or df1.empty) and (df2 is None or df2.empty):
        return pd.DataFrame()
    elif df1 is None or df1.empty:
        return df2.copy()
    elif df2 is None or df2.empty:
        return df1.copy()
    if 'Status Aceite' not in df1.columns:
        df1['Status Aceite'] = None
    if 'Status Aceite' not in df2.columns:
        df2['Status Aceite'] = None
    colunas_merge = ['Data', 'CPF/CNPJ Prestador', 'Razão Social/Nome do Prestador',
                     'Número', 'Valor do ISS', 'Valor dos Serviços', 'ISS Retido',
                     'Status Doc.', 'Status Aceite']
    colunas_merge_present = [c for c in colunas_merge if c in df1.columns and c in df2.columns]
    df = pd.merge(df1, df2, on=colunas_merge_present, how='outer', suffixes=('_fortaleza', '_vr'))
    origem_cols = [c for c in ['Origem_fortaleza', 'Origem_vr'] if c in df.columns]
    if len(origem_cols) == 2:
        df['Origem'] = df['Origem_fortaleza'].fillna(df['Origem_vr'])
        df.drop(columns=origem_cols, inplace=True)
    elif 'Origem_fortaleza' in df.columns:
        df.rename(columns={'Origem_fortaleza': 'Origem'}, inplace=True)
    elif 'Origem_vr' in df.columns:
        df.rename(columns={'Origem_vr': 'Origem'}, inplace=True)
    colunas_finais = ['Origem', 'Data', 'CPF/CNPJ Prestador', 'Razão Social/Nome do Prestador',
                      'Número', 'Valor do ISS', 'Valor dos Serviços', 'ISS Retido',
                      'Status Aceite', 'Status Doc.']
    existentes = [c for c in colunas_finais if c in df.columns]
    return df[existentes].copy()

def limpar_df_prefeitura(df):
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    if 'CPF/CNPJ Prestador' in df.columns:
        df['CPF/CNPJ Prestador'] = df['CPF/CNPJ Prestador'].apply(format_cnpj)
    if 'Status Doc.' in df.columns:
        df = df[df['Status Doc.'] != 'CANCELADA']
    if 'ISS Retido' in df.columns:
        df = df[~df['ISS Retido'].isin(['Não', 'NÃO'])]
    if 'Status Aceite' not in df.columns:
        df['Status Aceite'] = 'Não Informada'
    else:
        df = df[df['Status Aceite'] != 'Recusada']
    if 'Número' in df.columns:
        df['Número'] = df['Número'].astype(str).str.replace(r'\.0$', '', regex=True)
    if 'Valor do ISS' in df.columns:
        df['Valor do ISS'] = pd.to_numeric(df['Valor do ISS'], errors='coerce')
    return df

def limpar_df_financeiro(df):
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    if 'Histórico' in df.columns:
        df = df[~df['Histórico'].isin(['Saldo anterior'])]
        df = df[~df['Histórico'].str.startswith("PGTO.", na=False)]
    drop_cols = ['Numero_Credito', 'Código partida', 'Descrição partida', 'Contra-partida',
                 'Lote', 'Débito', 'Saldo', 'Cód Estab.', ' CNPJ Estab.',
                 'Operador', 'Data Geracao', 'Tipo (manual/automático)', 'Unnamed: 15']
    df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True, errors='ignore')
    if 'Histórico' in df.columns:
        df['Número'] = df['Histórico'].astype(str).str.extract(r'Doc\.\s*(\d+)', expand=False)
    if 'Documento' in df.columns:
        df['Número'] = df['Número'].fillna(df['Documento'].astype(str))
    if 'Número' in df.columns:
        df['Número'] = df['Número'].astype(str).str.replace(r'\.0$', '', regex=True)
    if 'Data' in df.columns:
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce', dayfirst=True)
        df['Data'] = df['Data'].dt.strftime('%d/%m/%Y')
    
    # Conversão de moeda para Crédito
    if 'Crédito' in df.columns:
        # Tenta converter direto se já for numérico (Excel), ou via parse se for texto
        if pd.api.types.is_numeric_dtype(df['Crédito']):
             df['Crédito'] = pd.to_numeric(df['Crédito'], errors='coerce')
        else:
             df['Crédito'] = parse_moeda_brasil_robusto(df['Crédito'])
    else:
        df['Crédito'] = np.nan
        
    return df

def criar_ids(df, numero_col, valor_col):
    if df is None or df.empty:
        return df
    df_temp = df.copy()
    if numero_col not in df_temp.columns or valor_col not in df_temp.columns:
        if numero_col in df_temp.columns:
            df_temp['ID'] = df_temp[numero_col].astype(str).str.replace(r'\.0$', '', regex=True)
        else:
            df_temp['ID'] = ""
        return df_temp
    
    # Tratamento para garantir que valor seja string limpa
    if df_temp[valor_col].dtype in ['float64', 'int64']:
         valor_str = df_temp[valor_col].astype(str).str.replace(r'\.0$', '', regex=True)
    else:
         valor_str = df_temp[valor_col].astype(str)

    numero_str = df_temp[numero_col].astype(str).str.replace(r'\.0$', '', regex=True)
    df_temp['ID'] = numero_str + valor_str
    return df_temp

def aplicar_validacao(df1, df2):
    if df1 is None:
        df1 = pd.DataFrame()
    if df2 is None:
        df2 = pd.DataFrame()
    df1_validado = df1.copy()
    df2_validado = df2.copy()
    if 'ID' in df1_validado.columns and 'ID' in df2_validado.columns:
        df1_validado['Status_Validacao'] = df1_validado['ID'].isin(df2_validado['ID']).map({True: 'Validado', False: 'Não Encontrado'})
        df2_validado['Status_Validacao'] = df2_validado['ID'].isin(df1_validado['ID']).map({True: 'Validado', False: 'Não Encontrado'})
    else:
        if 'ID' in df1_validado.columns:
            df1_validado['Status_Validacao'] = df1_validado['ID'].map(lambda x: 'Não Encontrado')
        else:
            df1_validado['Status_Validacao'] = np.nan
        if 'ID' in df2_validado.columns:
            df2_validado['Status_Validacao'] = df2_validado['ID'].map(lambda x: 'Não Encontrado')
        else:
            df2_validado['Status_Validacao'] = np.nan
    return df1_validado, df2_validado

def exportar_para_excel_bytes(df1, df2):
    output = BytesIO()

    # REGRA VR — garantir Status Aceite preenchido no output final
    if 'Status Aceite' in df1.columns:
        df1['Status Aceite'] = (
            df1['Status Aceite']
            .fillna('Não informada')
            .replace('', 'Não informada')
        )

    df1_export = df1.drop(
        columns=[col for col in ['ISS Retido', 'Status Doc.', 'ID'] if col in df1.columns],
        errors='ignore'
    )
    df2_export = df2.drop(
        columns=[col for col in ['ID'] if col in df2.columns],
        errors='ignore'
    )

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df1_export.to_excel(writer, sheet_name='Prefeitura', index=False)
        df2_export.to_excel(writer, sheet_name='Financeiro', index=False)

        ws1 = writer.sheets['Prefeitura']
        ws2 = writer.sheets['Financeiro']

        try:
            ws1.auto_filter.ref = ws1.dimensions
            ws2.auto_filter.ref = ws2.dimensions
        except Exception:
            pass

    output.seek(0)
    wb = openpyxl.load_workbook(output)

    fill_verde = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    for sheet_name in ['Prefeitura', 'Financeiro']:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        colunas_a_formatar = ['Valor do ISS', 'Valor dos Serviços', 'Crédito']
        colunas_idx = {
            cell.value: idx + 1
            for idx, cell in enumerate(ws[1])
            if cell.value in colunas_a_formatar
        }

        for col_name, col_idx in colunas_idx.items():
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                try:
                    cell.number_format = '#,##0.00'
                except Exception:
                    pass

        col_idx_val = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Status_Validacao":
                col_idx_val = idx
                break

        if col_idx_val:
            col_letter_val = openpyxl.utils.get_column_letter(col_idx_val)
            ws.conditional_formatting.add(
                f"{col_letter_val}2:{col_letter_val}{ws.max_row}",
                CellIsRule(
                    operator='equal',
                    formula=['"Validado"'],
                    fill=fill_verde
                )
            )

    out2 = BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2


def conciliar_notas(file_fortaleza=None, file_vr=None, file_razao=None, progress_callback=None):
    logs = []
    def p(pct, msg=None):
        if progress_callback:
            try:
                progress_callback(pct, msg)
            except Exception:
                pass
    p(5, "Iniciando leitura de arquivos.")
    if file_fortaleza is not None:
        try:
            df_fortaleza = preparar_dataframe_fortaleza(file_fortaleza)
        except Exception as e:
            df_fortaleza = pd.DataFrame()
            logs.append(f"Erro ao processar Fortaleza: {e}")
    else:
        df_fortaleza = pd.DataFrame()
        logs.append("Fortaleza não fornecido.")
    if file_vr is not None:
        try:
            df_vr = preparar_dataframe_vr(file_vr)
        except Exception as e:
            df_vr = pd.DataFrame()
            logs.append(f"Erro ao processar Volta Redonda: {e}")
    else:
        df_vr = pd.DataFrame()
        logs.append("Volta Redonda não fornecido.")
    p(40, "Unificando registros das Prefeituras.")
    df_unificado = unificar_dataframes(df_fortaleza, df_vr)
    
    # --- BLOCO CORRIGIDO DE LEITURA ---
    p(55, "Lendo arquivo Razão.")
    df_financeiro_raw = pd.DataFrame()
    if file_razao is not None:
        try:
            nome_arquivo = file_razao.name.lower()
            if hasattr(file_razao, "seek"):
                file_razao.seek(0)
            
            if nome_arquivo.endswith(('.xls', '.xlsx')):
                # Para Excel
                df_financeiro_raw = pd.read_excel(file_razao)
                # Garante limpeza da coluna Crédito se for lida como string
                if 'Crédito' in df_financeiro_raw.columns:
                     if not pd.api.types.is_numeric_dtype(df_financeiro_raw['Crédito']):
                        df_financeiro_raw['Crédito'] = parse_moeda_brasil_robusto(df_financeiro_raw['Crédito'])
            
            elif nome_arquivo.endswith('.csv'):
                # Para CSV (com prioridade no ;)
                df_financeiro_raw = carregar_arquivo_csv(file_razao)
            else:
                logs.append("Arquivo Razão em formato não suportado (apenas CSV, XLS, XLSX).")

        except Exception as e:
            df_financeiro_raw = pd.DataFrame()
            logs.append(f"Erro ao carregar Razão: {e}")
    else:
        df_financeiro_raw = pd.DataFrame()
        logs.append("Razão (financeiro) não fornecido.")
    # --- FIM DO BLOCO CORRIGIDO ---

    p(70, "Limpando dados.")
    df_prefeitura = limpar_df_prefeitura(df_unificado)
    df_financeiro = limpar_df_financeiro(df_financeiro_raw)
    p(82, "Gerando IDs de conciliação.")
    df_prefeitura = criar_ids(df_prefeitura, 'Número', 'Valor do ISS')
    df_financeiro = criar_ids(df_financeiro, 'Número', 'Crédito')
    p(92, "Aplicando validação cruzada.")
    df_prefeitura_valid, df_financeiro_valid = aplicar_validacao(df_prefeitura, df_financeiro)
    p(97, "Gerando arquivo Excel para download.")
    excel_buffer = exportar_para_excel_bytes(df_prefeitura_valid, df_financeiro_valid)
    p(100, "Concluído.")
    return df_prefeitura_valid, df_financeiro_valid, excel_buffer, logs

    # =========================================================
    # RESUMO DA CONCILIAÇÃO
    # =========================================================
    st.markdown("### 📊 Resumo da Conciliação")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("#### 🏛️ Prefeitura")

        total_pref = len(df_prefeitura_valid)
        validados_pref = (df_prefeitura_valid['Status_Validacao'] == 'Validado').sum()
        nao_encontrados_pref = (df_prefeitura_valid['Status_Validacao'] == 'Não Encontrado').sum()

        st.metric("Total de registros", total_pref)
        st.metric("✅ Validados", validados_pref)
        st.metric("❌ Não encontrados", nao_encontrados_pref)

    with col2:
        st.markdown("#### 💰 Financeiro")

        total_fin = len(df_financeiro_valid)
        validados_fin = (df_financeiro_valid['Status_Validacao'] == 'Validado').sum()
        nao_encontrados_fin = (df_financeiro_valid['Status_Validacao'] == 'Não Encontrado').sum()

        st.metric("Total de registros", total_fin)
        st.metric("✅ Validados", validados_fin)
        st.metric("❌ Não encontrados", nao_encontrados_fin)

def pagina_conciliacao_iss():

    colh1, colh2 = st.columns([4,1])
    with colh1:
        st.markdown('<div class="big-title">Conciliação do ISS Retido</div>', unsafe_allow_html=True)
        st.markdown('<div class="sub-title">Automação fiscal personalizada para LIV SAÚDE.</div>', unsafe_allow_html=True)

    st.markdown("---")

    st.markdown('<div class="panel">', unsafe_allow_html=True)
    st.subheader("Upload dos Documentos")

    col1, col2, col3 = st.columns(3)
    with col1:
        file_fortaleza = st.file_uploader("📄 NFS Fortaleza", type=["xlsx"])
    with col2:
        file_vr = st.file_uploader("📄 NFS Volta Redonda", type=["xls", "xlsx"])
    with col3:
        file_razao = st.file_uploader("📊 Razão Contábil", type=["csv", "xls", "xlsx"])

    st.markdown('</div>', unsafe_allow_html=True)

    if 'logs' not in st.session_state:
        st.session_state.logs = []
    if 'progress' not in st.session_state:
        st.session_state.progress = 0

    def progress_cb(pct, msg=None):
        st.session_state.progress = int(pct)
        if msg:
            st.session_state.logs.append(f"{pct}% - {msg}")

    if st.button("🚀 Processar"):
        st.session_state.logs = []
        st.session_state.progress = 0

        with st.spinner("Executando conciliação..."):
            df_pref, df_fin, excel_buf, logs = conciliar_notas(
                file_fortaleza,
                file_vr,
                file_razao,
                progress_callback=progress_cb
            )

            for l in logs:
                st.session_state.logs.append(l)

            st.success("Conciliação concluída!")
            
            # 1️⃣ Log de execução
            #st.subheader("📘 Log de Execução")
            #for l in st.session_state.logs[-200:]:
            #    st.write("•", l)

            # 2️⃣ Resumo da conciliação
            st.markdown("### 📊 Resumo da Conciliação")

            col1, col2 = st.columns(2)

            # Resumo da Prefeitura
            with col1:
                st.markdown("#### 🏛️ Prefeitura")

                total_pref = len(df_pref)
                validados_pref = (df_pref['Status_Validacao'] == 'Validado').sum()
                nao_encontrados_pref = (df_pref['Status_Validacao'] == 'Não Encontrado').sum()

                st.metric("Total de registros", total_pref)
                st.metric("✅ Validados", validados_pref)
                st.metric("❌ Não encontrados", nao_encontrados_pref)

            # Resumo do Financeiro
            with col2:
                st.markdown("#### 💰 Financeiro")

                total_fin = len(df_fin)
                validados_fin = (df_fin['Status_Validacao'] == 'Validado').sum()
                nao_encontrados_fin = (df_fin['Status_Validacao'] == 'Não Encontrado').sum()

                st.metric("Total de registros", total_fin)
                st.metric("✅ Validados", validados_fin)
                st.metric("❌ Não encontrados", nao_encontrados_fin)

            # 3️⃣ Botão para baixar planilha  
            if excel_buf:
                st.download_button(
                    "📥 Baixar Planilha Conciliada",
                    data=excel_buf.getvalue(),
                    file_name="Planilha Conciliada.xlsx"
                )

